#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para actualizar las pizzas del menu.json desde el Excel NAPOLI-PIZZAS.xlsx
"""

from openpyxl import load_workbook
import json
import re

# Leer el Excel
wb = load_workbook('NAPOLI-PIZZAS.xlsx')
ws = wb.active

# Leer pizzas del Excel
pizzas_excel = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:  # Si tiene nombre
        pizzas_excel.append({
            'nombre': row[0],
            'personal': row[1],
            'mediana': row[2],
            'familiar': row[3]
        })

print(f"✓ Leídas {len(pizzas_excel)} pizzas del Excel")

# Leer el menu.json actual
with open('src/data/menu.json', 'r', encoding='utf-8') as f:
    menu_data = json.load(f)

# Encontrar la sección de pizzas
pizza_section = None
for section in menu_data:
    if section['id'] == 'pizzas':
        pizza_section = section
        break

if not pizza_section:
    print("❌ No se encontró la sección de pizzas")
    exit(1)

# Encontrar la subcategoría de pizzas tradicionales
tradicionales_subcat = None
for subcat in pizza_section['subcategories']:
    if subcat['id'] == 'pizzas-tradicionales':
        tradicionales_subcat = subcat
        break

if not tradicionales_subcat:
    print("❌ No se encontró la subcategoría de pizzas tradicionales")
    exit(1)

# Encontrar la categoría de pizzas clásicas
pizzas_category = None
for cat in tradicionales_subcat['categories']:
    if cat['id'] == 'clasicas':
        pizzas_category = cat
        break

if not pizzas_category:
    print("❌ No se encontró la categoría de pizzas clásicas")
    exit(1)

# Generar ID a partir del nombre
def generar_id(nombre):
    # Convertir a minúsculas, quitar espacios y caracteres especiales
    id_base = nombre.lower().strip()
    id_base = re.sub(r'[^a-z0-9\s-]', '', id_base)
    id_base = re.sub(r'\s+', '-', id_base)
    return f"pizza-{id_base}"

# Generar descripción automática
def generar_descripcion(nombre):
    # Descripción genérica basada en el nombre
    ingredientes = nombre.lower()
    if 'queso' in ingredientes:
        base = "Deliciosa pizza con una mezcla especial de quesos"
    elif 'carne' in ingredientes:
        base = "Exquisita combinación de carnes selectas"
    elif 'vegetariana' in ingredientes or 'vegetales' in ingredientes:
        base = "Fresca selección de vegetales de temporada"
    elif 'hawaiana' in ingredientes:
        base = "Clásica combinación de jamón y piña"
    elif 'pollo' in ingredientes:
        base = "Jugoso pollo con ingredientes frescos"
    elif 'pepperoni' in ingredientes:
        base = "Tradicional pizza con pepperoni premium"
    else:
        base = "Pizza artesanal con ingredientes frescos"
    
    return f"{base}. Disponible en tres tamaños."

# Imagen genérica para todas las pizzas
IMAGEN_GENERICA = "https://pacciolo-legal-autos.s3.us-east-1.amazonaws.com/imagenes_proyectos/Napoli_Pizza_Generica.jpg"

# Convertir pizzas del Excel al formato del menu.json
nuevas_pizzas = []
for pizza in pizzas_excel:
    producto = {
        "id": generar_id(pizza['nombre']),
        "name": f"Pizza {pizza['nombre'].title()}",
        "price": pizza['mediana'],  # Precio base (Mediana)
        "description": generar_descripcion(pizza['nombre']),
        "image": IMAGEN_GENERICA,
        "attributes": ["Masa artesanal", "Ingredientes frescos"],
        "prepTime": "25-30 min",
        "sizePrices": {
            "Personal": pizza['personal'],
            "Mediana": pizza['mediana'],
            "Familiar": pizza['familiar']
        }
    }
    nuevas_pizzas.append(producto)

# Reemplazar las pizzas en la categoría
pizzas_category['products'] = nuevas_pizzas

print(f"✓ Generados {len(nuevas_pizzas)} productos de pizza")

# Guardar el menu.json actualizado
with open('src/data/menu.json', 'w', encoding='utf-8') as f:
    json.dump(menu_data, f, indent=2, ensure_ascii=False)

print(f"✅ menu.json actualizado exitosamente con {len(nuevas_pizzas)} pizzas")
print(f"   - Todas en la categoría 'Pizzas Tradicionales'")
print(f"   - Con 3 tamaños: Personal, Mediana, Familiar")
print(f"   - Descripciones automáticas generadas")
print(f"   - Imagen genérica asignada")
